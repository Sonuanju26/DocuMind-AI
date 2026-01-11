from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends
from sqlalchemy.orm import Session
from typing import List, Optional
import ollama
import json
from io import BytesIO
import PyPDF2
import docx
import mammoth
import openpyxl
import logging
import requests
import os
from ..database import get_db
from .. import crud

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

router = APIRouter()

def extract_text_from_file(file: UploadFile) -> str:
    """Extract text from various file types"""
    filename = file.filename.lower()
    file.file.seek(0)
    
    try:
        if filename.endswith('.txt'):
            content = file.file.read().decode('utf-8', errors='ignore')
            logger.info(f"Extracted {len(content)} characters from TXT file")
            return content
        
        elif filename.endswith('.pdf'):
            pdf_reader = PyPDF2.PdfReader(file.file)
            text = ""
            for page_num, page in enumerate(pdf_reader.pages):
                page_text = page.extract_text() or ""
                text += page_text
                logger.info(f"Extracted page {page_num + 1}/{len(pdf_reader.pages)}")
            logger.info(f"Total PDF text: {len(text)} characters")
            return text
        
        elif filename.endswith('.docx'):
            doc = docx.Document(BytesIO(file.file.read()))
            text = "\n".join([p.text for p in doc.paragraphs])
            logger.info(f"Extracted {len(text)} characters from DOCX")
            return text
        
        elif filename.endswith('.doc'):
            result = mammoth.extract_raw_text(BytesIO(file.file.read()))
            logger.info(f"Extracted {len(result.value)} characters from DOC")
            return result.value
        
        elif filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(BytesIO(file.file.read()), data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows():
                    values = [str(cell.value) if cell.value else "" for cell in row]
                    text += " ".join(values) + "\n"
            logger.info(f"Extracted {len(text)} characters from XLSX")
            return text
        
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {filename}")
    
    except Exception as e:
        logger.error(f"Error extracting text from {filename}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error extracting text: {str(e)}")

def check_ollama_connection():
    """Check if Ollama service is running and accessible"""
    ollama_host = os.getenv("OLLAMA_HOST", "http://localhost:11434")
    
    try:
        logger.info(f"Checking Ollama connection at {ollama_host}")
        response = requests.get(f"{ollama_host}/api/tags", timeout=5)
        
        if response.status_code == 200:
            models = response.json()
            logger.info(f"Ollama is running. Available models: {models}")
            return True
        else:
            logger.error(f"Ollama returned status code: {response.status_code}")
            return False
            
    except requests.exceptions.ConnectionError:
        logger.error(f"Cannot connect to Ollama at {ollama_host}")
        raise HTTPException(
            status_code=503,
            detail=f"Cannot connect to Ollama service at {ollama_host}. Please ensure Ollama is running. Start it with 'ollama serve' command."
        )
    except requests.exceptions.Timeout:
        logger.error("Ollama connection timeout")
        raise HTTPException(
            status_code=503,
            detail="Ollama service timeout. Please check if Ollama is running properly."
        )
    except Exception as e:
        logger.error(f"Error checking Ollama: {str(e)}")
        raise HTTPException(
            status_code=503,
            detail=f"Error connecting to Ollama: {str(e)}"
        )

def generate_summary_with_ollama(text: str, length: str, style: str, user_prompt: str = None) -> str:
    """Generate summary using Ollama Llama 3.2"""
    
    logger.info(f"Starting summarization - Length: {length}, Style: {style}")
    logger.info(f"Text length: {len(text)} characters")
    
    # Check Ollama connection first
    check_ollama_connection()
    
    length_tokens = {
        "short": "100-150 words",
        "medium": "250-300 words",
        "long": "500-600 words"
    }
    
    style_instructions = {
        "paragraph": "Write in clear, flowing paragraphs.",
        "bullet": "Use bullet points with • symbol. Start each point on a new line.",
        "flashcard": "Format as Q&A flashcards. Use 'Q:' and 'A:' prefixes.",
        "mindmap": "Create a hierarchical structure with main topics and subtopics using indentation.",
        "keypoints": "List the key points numbered 1, 2, 3, etc."
    }
    
    # Truncate text if too long
    max_text_length = 15000
    if len(text) > max_text_length:
        text = text[:max_text_length]
        logger.info(f"Text truncated to {max_text_length} characters")
    
    prompt = f"""Summarize the following text.
Length: {length_tokens.get(length, '250-300 words')}
Style: {style_instructions.get(style, 'Write in clear paragraphs.')}
{f'Additional instructions: {user_prompt}' if user_prompt else ''}

Text to summarize:
{text}

Summary:"""
    
    try:
        logger.info("Calling Ollama API to generate summary...")
        
        # Call Ollama with options
        response = ollama.generate(
            model='llama3.2:3b',
            prompt=prompt,
            options={
                'temperature': 0.7,
                'top_p': 0.9,
                'top_k': 40,
            }
        )
        
        logger.info("Ollama API response received successfully")
        
        if not response or 'response' not in response:
            logger.error("Ollama returned empty or invalid response")
            raise HTTPException(status_code=500, detail="Ollama returned empty response")
        
        summary = response['response'].strip()
        logger.info(f"Generated summary length: {len(summary)} characters")
        
        return summary
        
    except HTTPException:
        raise
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Ollama generation error: {error_msg}")
        
        # Check for specific errors
        if "model" in error_msg.lower() and "not found" in error_msg.lower():
            raise HTTPException(
                status_code=500, 
                detail="Model 'llama3.2:3b' not found. Please run: ollama pull llama3.2:3b"
            )
        elif "connection" in error_msg.lower():
            raise HTTPException(
                status_code=503,
                detail="Lost connection to Ollama service. Please check if Ollama is still running."
            )
        else:
            raise HTTPException(status_code=500, detail=f"Error generating summary: {error_msg}")

@router.post("/summarize")
async def summarize_files(
    files: List[UploadFile] = File(...),
    settings_json: Optional[str] = Form(None),
    user_id: Optional[int] = Form(None),
    db: Session = Depends(get_db)
):
    logger.info(f"Received summarization request for {len(files)} file(s)")
    
    settings = {"length": "medium", "style": "paragraph", "userQuery": ""}
    
    if settings_json:
        try:
            settings.update(json.loads(settings_json))
            logger.info(f"Settings: {settings}")
        except Exception as e:
            logger.warning(f"Error parsing settings: {e}")
    
    results = []
    
    for idx, file in enumerate(files):
        logger.info(f"Processing file {idx + 1}/{len(files)}: {file.filename}")
        
        try:
            # Extract text
            text = extract_text_from_file(file)
            
            if not text.strip():
                logger.warning(f"File {file.filename} is empty")
                results.append({
                    "fileName": file.filename,
                    "error": "File is empty or contains no readable text"
                })
                continue
            
            # Generate summary
            logger.info(f"Generating summary for {file.filename}")
            summary = generate_summary_with_ollama(
                text,
                settings.get('length', 'medium'),
                settings.get('style', 'paragraph'),
                settings.get('userQuery')
            )
            
            # Save to database
            if user_id:
                logger.info(f"Saving summary to database for user {user_id}")
                try:
                    crud.create_summary(
                        db,
                        user_id=user_id,
                        file_name=file.filename,
                        original_text=text,
                        summary_text=summary,
                        length=settings.get('length'),
                        style=settings.get('style'),
                        user_prompt=settings.get('userQuery')
                    )
                    logger.info("Summary saved to database successfully")
                except Exception as db_error:
                    logger.error(f"Database error: {str(db_error)}")
                    # Don't fail the request if DB save fails
            
            results.append({
                "fileName": file.filename,
                "summary": summary
            })
            
            logger.info(f"Successfully processed {file.filename}")
            
        except HTTPException as he:
            logger.error(f"HTTP error processing {file.filename}: {he.detail}")
            results.append({
                "fileName": file.filename,
                "error": he.detail
            })
        except Exception as e:
            logger.error(f"Unexpected error processing {file.filename}: {str(e)}")
            results.append({
                "fileName": file.filename,
                "error": f"Unexpected error: {str(e)}"
            })
    
    logger.info(f"Completed processing {len(files)} file(s). Successful: {sum(1 for r in results if 'summary' in r)}")
    return results